"""
ChatKit Backend - Custom Chat API with multiple AI models support.
"""
from __future__ import annotations

import asyncio
import base64
import json
import logging
import os
import re
import httpx
from datetime import datetime
from typing import Any, AsyncIterator, Annotated
from contextlib import asynccontextmanager
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

from chatkit.agents import AgentContext
from fastapi import FastAPI, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openai import AsyncOpenAI
from agents import RunConfig, Runner, function_tool, RunContextWrapper, WebSearchTool, Agent, OpenAIConversationsSession
from agents.exceptions import ModelBehaviorError

# OpenAI Async Client for Assistants API
openai_client = AsyncOpenAI()
from agents.model_settings import ModelSettings
from starlette.responses import JSONResponse

from firebase_service import get_firebase_service
from claude_document_agent import get_claude_agent
import google.generativeai as genai

logger = logging.getLogger(__name__)

# API Keys for AI models
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "AIzaSyCucua6TsC4iR5qEaFHz710kAk_U_rO_AQ")
DEEPSEEK_API_KEY = os.environ.get("DEEPSEEK_API_KEY", "sk-674ee5a599124755b6217eac65505e59")

# Configure Gemini
genai.configure(api_key=GEMINI_API_KEY)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


# ─────────────────────────────────────────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────────────────────────────────────────

CLAUDE_WORKSPACE_DIR = os.environ.get("CLAUDE_WORKSPACE_DIR", "./claude_workspace")

# Create workspace directory for Claude document processing
os.makedirs(CLAUDE_WORKSPACE_DIR, exist_ok=True)
logger.info(f"Claude workspace directory: {CLAUDE_WORKSPACE_DIR}")

# ─────────────────────────────────────────────────────────────────────────────
# Global HTTP Client with Connection Pooling (OPTIMIZATION)
# ─────────────────────────────────────────────────────────────────────────────
# This dramatically reduces latency for external API calls by reusing connections
from typing import Optional, Dict, Any, Tuple
_http_client: Optional[httpx.AsyncClient] = None

def get_http_client() -> httpx.AsyncClient:
    """Get global HTTP client with connection pooling."""
    global _http_client
    if _http_client is None:
        _http_client = httpx.AsyncClient(
            timeout=30.0,
            limits=httpx.Limits(max_connections=100, max_keepalive_connections=20),
            http2=True  # Enable HTTP/2 for better performance
        )
        logger.info("✅ Created global HTTP client with connection pooling")
    return _http_client


# ─────────────────────────────────────────────────────────────────────────────
# Weather/AQI Cache (OPTIMIZATION)
# ─────────────────────────────────────────────────────────────────────────────
# Cache weather and AQI data to reduce API calls and improve response time

class SimpleCache:
    """Simple in-memory cache with TTL."""
    def __init__(self, ttl_seconds: int = 1800):  # 30 minutes default
        self._cache: Dict[str, Tuple[Any, float]] = {}
        self._ttl = ttl_seconds

    def get(self, key: str) -> Optional[Any]:
        """Get cached value if not expired."""
        import time
        if key in self._cache:
            value, timestamp = self._cache[key]
            if time.time() - timestamp < self._ttl:
                logger.info(f"📦 Cache HIT for: {key}")
                return value
            else:
                # Expired, remove
                del self._cache[key]
        return None

    def set(self, key: str, value: Any) -> None:
        """Set cache value with current timestamp."""
        import time
        self._cache[key] = (value, time.time())
        logger.info(f"📦 Cache SET for: {key}")

    def clear(self) -> None:
        """Clear all cached data."""
        self._cache.clear()

# Global caches
_weather_cache = SimpleCache(ttl_seconds=1800)  # 30 min for weather
_aqi_cache = SimpleCache(ttl_seconds=3600)  # 60 min for AQI (changes less frequently)


# ─────────────────────────────────────────────────────────────────────────────
# Tools
# ─────────────────────────────────────────────────────────────────────────────

# Создаем базовый web search tool - это встроенный OpenAI tool
web_search_tool = WebSearchTool(
    search_context_size="medium",
    user_location={
        "city": "Tashkent",
        "country": "UZ",
        "region": "Tashkent",
        "type": "approximate"
    }
)


# ─────────────────────────────────────────────────────────────────────────────
# FastAPI Application
# ─────────────────────────────────────────────────────────────────────────────


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Application lifespan manager - handles startup and shutdown."""
    # Startup
    logger.info("=" * 80)
    logger.info("🚀 ChatKit Backend - Starting up")
    logger.info("=" * 80)

    # Initialize Firebase service
    try:
        firebase = get_firebase_service()
        if firebase._initialized:
            logger.info("=" * 80)
            logger.info("✅ All services initialized successfully")
            logger.info("=" * 80)
        else:
            logger.info("=" * 80)
            logger.info("⚠️  Running WITHOUT Firebase (messages won't be saved)")
            logger.info("=" * 80)
    except Exception as e:
        logger.error(f"❌ Failed to initialize Firebase service: {e}")
        logger.error("=" * 80)

    yield

    # Shutdown
    logger.info("=" * 80)
    logger.info("👋 ChatKit Backend - Shutting down")

    # Close global HTTP client
    global _http_client
    if _http_client:
        await _http_client.aclose()
        _http_client = None
        logger.info("✅ Closed global HTTP client")

    logger.info("=" * 80)


app = FastAPI(title="ChatKit Backend", version="2.0.0", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://click.ailab.uz",
        "https://react-firebase-chat-beta-steel.vercel.app",
        "https://dist-telegram.vercel.app",
        "https://clickapp-bo7d.vercel.app",
        "http://localhost:3000",
        "http://localhost:5173",
        "http://localhost:5174",
    ],
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS", "PATCH"],
    allow_headers=["*"],
    expose_headers=["*"],
    max_age=3600,
)


@app.get("/")
def read_root():
    return {
        "service": "ChatKit Backend",
        "status": "running",
        "version": "2.0.0",
        "endpoints": ["/health", "/api/custom-chat", "/api/aqi-widget"],
    }


@app.get("/health")
def health_check():
    return {"status": "healthy", "service": "chatkit-backend", "version": "2.0.0"}


# Custom tools for custom-chat endpoint
@function_tool
async def get_weather_simple(
    ctx: RunContextWrapper[AgentContext],
    location: Annotated[str, "City name or location to get weather for"]
) -> str:
    """Get current weather information for a specified location with interactive widget."""
    logger.info(f"=" * 60)
    logger.info(f"🌤️ [WEATHER TOOL] Called with location parameter: '{location}'")
    logger.info(f"🌤️ [WEATHER TOOL] Location type: {type(location)}")
    logger.info(f"=" * 60)

    # Скелетоны отключены - виджеты показываются сразу когда готовы
    logger.info(f"🔄 Weather widget loading started for {location}")

    # Check cache first
    cache_key = f"weather_{location.lower().strip()}"
    cached_data = _weather_cache.get(cache_key)
    if cached_data:
        logger.info(f"🌤️ [WEATHER TOOL] Cache HIT for {location}")
        # Add widget from cache to context
        if not hasattr(ctx.context, '_custom_widgets'):
            ctx.context._custom_widgets = []
        ctx.context._custom_widgets.append(cached_data['widget'])
        return cached_data['response']

    logger.info(f"🌤️ [WEATHER TOOL] Cache MISS for {location}")

    try:
        client = get_http_client()
        # Get location coordinates - try Russian first, then English
        geo_response = await client.get(
            "https://geocoding-api.open-meteo.com/v1/search",
            params={"name": location, "count": 5, "language": "ru", "format": "json"},
            timeout=10.0
        )

        if geo_response.status_code != 200:
            return f"Could not find location: {location}"

        geo_data = geo_response.json()
        logger.info(f"🌤️ [WEATHER TOOL] Geocoding results count: {len(geo_data.get('results', []))}")
        if not geo_data.get("results"):
            logger.warning(f"🌤️ [WEATHER TOOL] No results found for location: {location}")
            return f"Location not found: {location}"

        result = geo_data["results"][0]
        lat = result["latitude"]
        lon = result["longitude"]
        location_name = result["name"]
        country = result.get("country", "")
        logger.info(f"🌤️ [WEATHER TOOL] Found: {location_name}, {country} ({lat}, {lon})")

        # Get weather data with hourly forecast for 7 days
        weather_response = await client.get(
            "https://api.open-meteo.com/v1/forecast",
            params={
                "latitude": lat,
                "longitude": lon,
                "current": "temperature_2m,relative_humidity_2m,weather_code,wind_speed_10m",
                "hourly": "temperature_2m,weather_code",
                "daily": "weather_code,temperature_2m_max,temperature_2m_min",
                "timezone": "auto",
                "forecast_days": 7,
                "forecast_hours": 168  # 7 days * 24 hours
            },
            timeout=10.0
        )
        logger.info(f"🌤️ [WEATHER TOOL] Weather API response status: {weather_response.status_code}")

        if weather_response.status_code != 200:
            logger.error(f"🌤️ [WEATHER TOOL] Weather API error: {weather_response.status_code}")
            return f"Could not fetch weather data for {location}"

        weather_data = weather_response.json()
        logger.info(f"🌤️ [WEATHER TOOL] Weather data keys: {list(weather_data.keys())}")
        current = weather_data["current"]
        daily = weather_data.get("daily", {})
        hourly = weather_data.get("hourly", {})

        weather_codes = {
            0: "Clear", 1: "Mainly Clear", 2: "Partly Cloudy", 3: "Cloudy",
            45: "Foggy", 48: "Foggy", 51: "Light Drizzle",
            61: "Light Rain", 63: "Moderate Rain", 65: "Heavy Rain",
            71: "Light Snow", 73: "Moderate Snow", 75: "Heavy Snow",
            80: "Rain Showers", 95: "Thunderstorm"
        }

        temperature = current.get("temperature_2m", 0)
        humidity = current.get("relative_humidity_2m", 0)
        wind_speed = current.get("wind_speed_10m", 0)
        weather_code = current.get("weather_code", 0)
        condition = weather_codes.get(weather_code, "Unknown")
        logger.info(f"🌤️ [WEATHER TOOL] Current weather: {temperature}°C, {condition}")

        # Build daily forecast with hourly data for each day
        forecast = []
        if daily and "time" in daily:
            for i in range(min(7, len(daily["time"]))):
                date_str = daily["time"][i]
                date_obj = datetime.fromisoformat(date_str)
                day_name = date_obj.strftime("%a")

                # Get hourly data for this day
                day_hourly = []
                if hourly and "time" in hourly:
                    for h_idx, h_time in enumerate(hourly["time"]):
                        h_date = datetime.fromisoformat(h_time)
                        if h_date.date() == date_obj.date():
                            day_hourly.append({
                                "time": h_date.strftime("%H:%M"),
                                "hour": h_date.hour,
                                "temp": hourly["temperature_2m"][h_idx],
                                "weather_code": hourly.get("weather_code", [0] * len(hourly["time"]))[h_idx]
                            })

                forecast.append({
                    "day": day_name,
                    "date": date_str,
                    "temp_max": daily["temperature_2m_max"][i],
                    "temp_min": daily["temperature_2m_min"][i],
                    "condition": weather_codes.get(daily["weather_code"][i], "Unknown"),
                    "weather_code": daily["weather_code"][i],
                    "hourly": day_hourly
                })

        # Create widget data для frontend
        full_location = f"{location_name}, {country}" if country else location_name
        logger.info(f"🌤️ [WEATHER TOOL] Creating widget with location: '{full_location}'")
        logger.info(f"🌤️ [WEATHER TOOL] location_name from API: '{location_name}', country: '{country}'")

        widget_data = {
            "_widget_type": "weather",
            "location": full_location,
            "temperature": round(temperature, 1),
            "condition": condition,
            "humidity": humidity,
            "wind_speed": round(wind_speed, 1),
            "forecast": forecast,
            "weather_code": weather_code
        }

        logger.info(f"🌤️ [WEATHER TOOL] Widget data keys: {list(widget_data.keys())}")
        logger.info(f"🌤️ [WEATHER TOOL] Widget location value: '{widget_data.get('location')}'")

        # Сохраняем widget в контексте для извлечения в streaming loop
        if not hasattr(ctx.context, '_custom_widgets'):
            ctx.context._custom_widgets = []
        ctx.context._custom_widgets.append(widget_data)
        logger.info(f"🌤️ [WEATHER TOOL] Widget added to context. Total widgets: {len(ctx.context._custom_widgets)}")

        logger.info(f"✅ Widget data prepared for {location_name}")

        # Return text description that LLM can see (while widget is shown separately)
        logger.info(f"🌤️ [WEATHER TOOL] Returning weather data for LLM")

        # Build forecast summary for next 3 days
        forecast_summary = ""
        if len(forecast) > 1:
            forecast_summary = "\n\nПрогноз на ближайшие дни:"
            for day_data in forecast[1:4]:  # Next 3 days
                forecast_summary += f"\n- {day_data['day']}: {day_data['condition']}, {round(day_data['temp_max'])}°C / {round(day_data['temp_min'])}°C"

        response_text = f"""Отправлен интерактивный виджет с данными о погоде для **{full_location}**:

- **Текущая температура:** {round(temperature, 1)}°C
- **Условия:** {condition}
- **Влажность:** {humidity}%
- **Скорость ветра:** {round(wind_speed, 1)} км/ч
{forecast_summary}"""

        # Save to cache
        _weather_cache.set(cache_key, {
            'widget': widget_data,
            'response': response_text
        })
        logger.info(f"🌤️ [WEATHER TOOL] Cached weather data for {location}")

        return response_text

    except Exception as e:
        logger.error(f"Weather fetch error: {e}")
        return f"Error fetching weather data: {str(e)}"


# IQAir API Key
IQAIR_API_KEY = os.environ.get("IQAIR_API_KEY", "9dab5d99-05fc-4359-bdaf-498590da28b4")


@function_tool
async def get_aqi_simple(
    ctx: RunContextWrapper[AgentContext],
    location: Annotated[str, "City name to get air quality index for"] = "Ташкент"
) -> str:
    """Get current Air Quality Index (AQI) information for a specified location with interactive widget using IQAir API."""
    logger.info(f"=" * 60)
    logger.info(f"🌫️ [AQI TOOL - IQAir] Called with location parameter: '{location}'")
    logger.info(f"=" * 60)

    # Check cache first
    cache_key = f"aqi_{location.lower().strip()}"
    cached_data = _aqi_cache.get(cache_key)
    if cached_data:
        logger.info(f"🌫️ [AQI TOOL] Cache HIT for {location}")
        # Add widget from cache to context
        if not hasattr(ctx.context, '_custom_widgets'):
            ctx.context._custom_widgets = []
        ctx.context._custom_widgets.append(cached_data['widget'])
        return cached_data['response']

    logger.info(f"🌫️ [AQI TOOL] Cache MISS for {location}")

    try:
        client = get_http_client()
        # First get coordinates for the location using geocoding
        geo_response = await client.get(
            "https://geocoding-api.open-meteo.com/v1/search",
            params={"name": location, "count": 1, "language": "ru", "format": "json"},
            timeout=10.0
        )

        if geo_response.status_code != 200:
            return f"Не удалось найти местоположение: {location}"

        geo_data = geo_response.json()
        if not geo_data.get("results"):
            return f"Местоположение не найдено: {location}"

        result = geo_data["results"][0]
        lat = result["latitude"]
        lon = result["longitude"]
        location_name = result["name"]

        logger.info(f"🌫️ [AQI TOOL] Coordinates: {lat}, {lon} for {location_name}")

        # Get air quality data from IQAir API
        iqair_response = await client.get(
            "https://api.airvisual.com/v2/nearest_city",
            params={
                "lat": lat,
                "lon": lon,
                "key": IQAIR_API_KEY
            },
            timeout=15.0
        )

        logger.info(f"🌫️ [AQI TOOL] IQAir response status: {iqair_response.status_code}")

        if iqair_response.status_code != 200:
            logger.error(f"IQAir API error: {iqair_response.text}")
            return f"Не удалось получить данные о качестве воздуха для {location}"

        iqair_data = iqair_response.json()
        logger.info(f"🌫️ [AQI TOOL] IQAir data: {iqair_data}")

        if iqair_data.get("status") != "success":
            return f"Ошибка IQAir API: {iqair_data.get('data', {}).get('message', 'Unknown error')}"

        data = iqair_data.get("data", {})
        current = data.get("current", {})
        pollution = current.get("pollution", {})
        weather = current.get("weather", {})

        # Get city name from IQAir or use geocoded name
        city_name = data.get("city", location_name)

        # Get AQI value (US AQI standard)
        aqi_value = int(pollution.get("aqius", 0))

        # Determine category based on US AQI
        if aqi_value <= 50:
            category = "Хорошо"
        elif aqi_value <= 100:
            category = "Умеренно"
        elif aqi_value <= 150:
            category = "Вредно для чувствительных"
        elif aqi_value <= 200:
            category = "Вредно"
        elif aqi_value <= 300:
            category = "Очень вредно"
        else:
            category = "Опасно"

        # Get main pollutant from IQAir
        main_pollutant_code = pollution.get("mainus", "p2")
        pollutant_map = {
            "p2": "PM2.5",
            "p1": "PM10",
            "o3": "O3",
            "n2": "NO2",
            "s2": "SO2",
            "co": "CO"
        }
        pollutant = pollutant_map.get(main_pollutant_code, "PM2.5")

        # Get weather data from IQAir response
        temp = f"{round(weather.get('tp', 0))}°C"
        humidity = f"{round(weather.get('hu', 0))}%"
        wind_speed = round(weather.get('ws', 0) * 3.6)  # Convert m/s to km/h
        wind = f"{wind_speed} км/ч"

        # Get current time
        now = datetime.now()
        updated = now.strftime("%H:%M, %d %B")

        # Create widget data for frontend
        logger.info(f"🌫️ [AQI TOOL] Creating widget for: {city_name}, AQI: {aqi_value}")

        widget_data = {
            "_widget_type": "aqi",
            "city": city_name,
            "aqi": aqi_value,
            "category": category,
            "scale": "US AQI",
            "pollutant": pollutant,
            "temp": temp,
            "humidity": humidity,
            "wind": wind,
            "updated": updated
        }

        # Save widget in context for extraction in streaming loop
        if not hasattr(ctx.context, '_custom_widgets'):
            ctx.context._custom_widgets = []
        ctx.context._custom_widgets.append(widget_data)

        logger.info(f"✅ AQI Widget (IQAir) prepared for {city_name}: AQI {aqi_value}")

        # Return text description that LLM can see
        response_text = f"""Отправлен интерактивный виджет с данными о качестве воздуха для **{city_name}**:

- **AQI:** {aqi_value} ({category})
- **Доминирующий загрязнитель:** {pollutant}
- **Температура:** {temp}
- **Влажность:** {humidity}
- **Ветер:** {wind}
- **Обновлено:** {updated}

*Данные предоставлены IQAir*"""

        # Save to cache
        _aqi_cache.set(cache_key, {
            'widget': widget_data,
            'response': response_text
        })
        logger.info(f"🌫️ [AQI TOOL] Cached AQI data for {location}")

        return response_text

    except Exception as e:
        logger.error(f"AQI fetch error (IQAir): {e}", exc_info=True)
        return f"Ошибка при получении данных о качестве воздуха: {str(e)}"


# ─────────────────────────────────────────────────────────────────────────────
# Handoff Tools - Gemini and DeepSeek
# ─────────────────────────────────────────────────────────────────────────────

@function_tool
async def ask_gemini(
    ctx: RunContextWrapper[AgentContext],
    query: Annotated[str, "The question or task to send to Gemini 2.5 Flash"],
    context: Annotated[str, "Additional context or background information"] = ""
) -> AsyncIterator[str]:
    """
    Handoff to Gemini 2.5 Flash model for advanced reasoning, analysis, or when user explicitly requests Gemini.
    Use this tool when:
    - User asks for Gemini specifically
    - Task requires advanced reasoning capabilities
    - Complex analysis or code review is needed
    - Multi-step reasoning is required
    """
    try:
        logger.info(f"🔮 [GEMINI HANDOFF] Calling Gemini 2.5 Flash with query: {query[:100]}...")

        # Build the prompt with context
        full_prompt = query
        if context:
            full_prompt = f"Context: {context}\n\nTask: {query}"

        # Yield header
        yield "**Ответ от Gemini 2.5 Flash:**\n\n"

        # Call Gemini API with streaming
        # Using stable version for production reliability
        model = genai.GenerativeModel('gemini-2.5-flash')
        response = await model.generate_content_async(
            full_prompt,
            generation_config=genai.types.GenerationConfig(
                max_output_tokens=4096,
                temperature=0.7,
            ),
            stream=True
        )

        # Stream chunks
        chunk_count = 0
        async for chunk in response:
            if chunk.text:
                chunk_count += 1
                yield chunk.text

        logger.info(f"✅ [GEMINI HANDOFF] Streamed {chunk_count} chunks")

    except Exception as e:
        logger.error(f"❌ [GEMINI HANDOFF] Error: {e}", exc_info=True)
        yield f"Ошибка при обращении к Gemini: {str(e)}"


@function_tool
async def ask_deepseek(
    ctx: RunContextWrapper[AgentContext],
    query: Annotated[str, "The question or task to send to DeepSeek V3"],
    context: Annotated[str, "Additional context or background information"] = ""
) -> AsyncIterator[str]:
    """
    Handoff to DeepSeek V3 model for specialized tasks, coding, or when user explicitly requests DeepSeek.
    Use this tool when:
    - User asks for DeepSeek specifically
    - Task involves complex coding or programming
    - Technical problem-solving is needed
    - Mathematical or scientific analysis is required
    """
    try:
        logger.info(f"🔮 [DEEPSEEK HANDOFF] Calling DeepSeek V3 with query: {query[:100]}...")

        # Build the prompt with context
        messages = []
        if context:
            messages.append({"role": "system", "content": f"Context: {context}"})
        messages.append({"role": "user", "content": query})

        # Yield header
        yield "**Ответ от DeepSeek V3:**\n\n"

        # Call DeepSeek API with streaming (OpenAI-compatible)
        async with httpx.AsyncClient(timeout=120.0) as client:
            async with client.stream(
                "POST",
                "https://api.deepseek.com/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
                    "Content-Type": "application/json"
                },
                json={
                    "model": "deepseek-chat",
                    "messages": messages,
                    "max_tokens": 4096,
                    "temperature": 0.7,
                    "stream": True
                }
            ) as response:
                if response.status_code != 200:
                    error_text = await response.aread()
                    logger.error(f"DeepSeek API error: {response.status_code} - {error_text}")
                    yield f"Ошибка DeepSeek API: {response.status_code}"
                    return

                chunk_count = 0
                async for line in response.aiter_lines():
                    if line.startswith("data: "):
                        data_str = line[6:]
                        if data_str == "[DONE]":
                            break
                        try:
                            data = json.loads(data_str)
                            if "choices" in data and len(data["choices"]) > 0:
                                delta = data["choices"][0].get("delta", {})
                                content = delta.get("content", "")
                                if content:
                                    chunk_count += 1
                                    yield content
                        except json.JSONDecodeError:
                            pass

                logger.info(f"✅ [DEEPSEEK HANDOFF] Streamed {chunk_count} chunks")

    except Exception as e:
        logger.error(f"❌ [DEEPSEEK HANDOFF] Error: {e}", exc_info=True)
        yield f"Ошибка при обращении к DeepSeek: {str(e)}"


@function_tool
def send_reaction_gif(
    ctx: RunContextWrapper[AgentContext],
    mood: Annotated[str, "The mood/emotion for the GIF: happy, celebration, thinking, confused, excited, sad, love, funny, greeting, thumbs_up, applause, mind_blown"],
    search_query: Annotated[str, "Optional specific search query for the GIF"] = ""
) -> str:
    """
    Send a reaction GIF to make the conversation more engaging and human-like.
    Use this SPARINGLY - only for:
    - Greetings (привет, здравствуй)
    - Celebrations (поздравления, успехи)
    - Jokes and humor
    - Expressing empathy
    - Special occasions

    DO NOT use for:
    - Every message
    - Serious/technical questions
    - When user seems in a hurry

    Limit: Maximum 1 GIF per 3-4 messages.
    """
    import random

    logger.info(f"🎬 [GIF] Mood: {mood}, Query: {search_query}")

    # Predefined GIF URLs by mood (using Giphy direct URLs - these are stable)
    gif_library = {
        "happy": [
            "https://media.giphy.com/media/l0MYGb1LuZ3n7dRnO/giphy.gif",
            "https://media.giphy.com/media/5GoVLqeAOo6PK/giphy.gif",
            "https://media.giphy.com/media/l46Cy1rHbQ92uuLXa/giphy.gif",
        ],
        "celebration": [
            "https://media.giphy.com/media/g9582DNuQppxC/giphy.gif",
            "https://media.giphy.com/media/artj92V8o75VPL7AeQ/giphy.gif",
            "https://media.giphy.com/media/3oz8xAFtqoOUUrsh7W/giphy.gif",
        ],
        "thinking": [
            "https://media.giphy.com/media/a5viI92PAF89q/giphy.gif",
            "https://media.giphy.com/media/3o7TKTDn976rzVgky4/giphy.gif",
            "https://media.giphy.com/media/CaiVJuZGvR8HK/giphy.gif",
        ],
        "excited": [
            "https://media.giphy.com/media/5VKbvrjxpVJCM/giphy.gif",
            "https://media.giphy.com/media/l0MYt5jPR6QX5pnqM/giphy.gif",
            "https://media.giphy.com/media/IwAZ6dvvvaTtdI8SD5/giphy.gif",
        ],
        "greeting": [
            "https://media.giphy.com/media/xT9IgG50Fb7Mi0prBC/giphy.gif",
            "https://media.giphy.com/media/bcKmIWkUMCjVm/giphy.gif",
            "https://media.giphy.com/media/dzaUX7CAG0Ihi/giphy.gif",
        ],
        "thumbs_up": [
            "https://media.giphy.com/media/111ebonMs90YLu/giphy.gif",
            "https://media.giphy.com/media/l41lUJ1YoZB1lHVPG/giphy.gif",
            "https://media.giphy.com/media/9xt1MUZqkneFiWrAAD/giphy.gif",
        ],
        "applause": [
            "https://media.giphy.com/media/nbvFVPiEiJH6JOGIok/giphy.gif",
            "https://media.giphy.com/media/l4q8gHsCDRGTR0MfK/giphy.gif",
            "https://media.giphy.com/media/fnK0jeA8vIh2QLq3IZ/giphy.gif",
        ],
        "love": [
            "https://media.giphy.com/media/l0HlvtIPzPdt2usKs/giphy.gif",
            "https://media.giphy.com/media/26BRv0ThflsHCqDrG/giphy.gif",
            "https://media.giphy.com/media/3oEjHV0z8S7WM4MwnK/giphy.gif",
        ],
        "funny": [
            "https://media.giphy.com/media/10JhviFuU2gWD6/giphy.gif",
            "https://media.giphy.com/media/3o6Zt4HU9uwXmXSAuI/giphy.gif",
            "https://media.giphy.com/media/l0MYryZTmQgvHI5sA/giphy.gif",
        ],
        "mind_blown": [
            "https://media.giphy.com/media/xT0xeJpnrWC4XWblEk/giphy.gif",
            "https://media.giphy.com/media/26ufdipQqU2lhNA4g/giphy.gif",
            "https://media.giphy.com/media/OK27wINdQS5YQ/giphy.gif",
        ],
        "sad": [
            "https://media.giphy.com/media/OPU6wzx8JrHna/giphy.gif",
            "https://media.giphy.com/media/k61nOBRRBMxva/giphy.gif",
        ],
        "confused": [
            "https://media.giphy.com/media/3o7btPCcdNniyf0ArS/giphy.gif",
            "https://media.giphy.com/media/WRQBXSCnEFJIuxktnw/giphy.gif",
        ],
    }

    # Get GIFs for the mood
    gifs = gif_library.get(mood, gif_library.get("happy", []))
    if not gifs:
        return ""  # No GIF to send

    # Select random GIF
    gif_url = random.choice(gifs)

    # Store GIF in context for frontend
    if not hasattr(ctx.context, '_reaction_gif'):
        ctx.context._reaction_gif = None

    ctx.context._reaction_gif = {
        "_gif_type": "reaction",
        "url": gif_url,
        "mood": mood
    }

    logger.info(f"🎬 [GIF] Selected: {gif_url}")
    return f"[GIF отправлен: {mood}]"


@function_tool
def create_document(
    ctx: RunContextWrapper[AgentContext],
    document_type: Annotated[str, "Type of document: 'presentation', 'excel', or 'word'"],
    task: Annotated[str, "Detailed description of what to create"],
    title: Annotated[str, "Title or name for the document"] = ""
) -> str:
    """
    Create a document (presentation, Excel table, or Word document).
    Use this tool when user asks to create:
    - Presentations (презентация, слайды, ppt)
    - Excel spreadsheets (таблица, excel, xlsx)
    - Word documents (документ, word, docx, договор, отчёт, письмо)

    The document will be created and sent to the user for download.
    """
    logger.info(f"📄 [CREATE_DOCUMENT] Type: {document_type}, Task: {task[:100]}...")

    # Store document request in context for frontend handling
    if not hasattr(ctx.context, '_document_request'):
        ctx.context._document_request = None

    ctx.context._document_request = {
        "_document_type": document_type,
        "task": task,
        "title": title or "document"
    }

    # Return confirmation message
    doc_type_names = {
        "presentation": "презентацию",
        "excel": "таблицу Excel",
        "word": "документ Word"
    }
    doc_name = doc_type_names.get(document_type, "документ")

    return f"Начинаю создание: {doc_name}. Пожалуйста, подождите - файл будет отправлен в чат."


@app.post("/api/custom-chat")
async def custom_chat_endpoint(
    request: Request
):
    """
    Custom chat endpoint using OpenAI Agents SDK with OpenAIConversationsSession.
    Supports text messages and image/document uploads.
    Uses Claude Haiku for document/image processing, GPT for conversation.
    Conversation history is stored in OpenAI cloud via OpenAIConversationsSession.
    """
    logger.info(f"=" * 80)
    logger.info(f"📨 [CUSTOM-CHAT] New request received (Agents SDK + OpenAIConversationsSession)")
    logger.info(f"=" * 80)

    try:
        content_type = request.headers.get("content-type", "")
        logger.info(f"📨 [CUSTOM-CHAT] Content-Type: {content_type}")

        # Parse request data
        conversation_id = None
        if "multipart/form-data" in content_type:
            form = await request.form()
            user_message = form.get("message", "")
            user_id = form.get("user_id")
            conversation_id = form.get("conversation_id")  # Get conversation_id from form
            selected_model = form.get("model", "gpt-5")  # Get selected model

            # Process uploaded files
            uploaded_files = []
            files_in_form = form.getlist("files")
            for file in files_in_form:
                if hasattr(file, 'file'):
                    file_bytes = await file.read()
                    uploaded_files.append({
                        "filename": file.filename,
                        "content_type": file.content_type or "application/octet-stream",
                        "bytes": file_bytes
                    })
                    logger.info(f"Received file: {file.filename}, size: {len(file_bytes)} bytes, type: {file.content_type}")
        else:
            data = await request.json()
            user_message = data.get("message", "")
            user_id = data.get("user_id")
            conversation_id = data.get("conversation_id")  # Get conversation_id from JSON
            selected_model = data.get("model", "gpt-5")  # Get selected model
            uploaded_files = []

        logger.info(f"📨 [CUSTOM-CHAT] User message: '{user_message[:200] if user_message else 'None'}...'")
        logger.info(f"📨 [CUSTOM-CHAT] User ID: {user_id}")
        logger.info(f"📨 [CUSTOM-CHAT] Conversation ID: {conversation_id}")
        logger.info(f"📨 [CUSTOM-CHAT] Selected model: {selected_model}")
        logger.info(f"📨 [CUSTOM-CHAT] Uploaded files: {len(uploaded_files)}")

        if not user_message and not uploaded_files:
            return JSONResponse({"error": "Message is required"}, status_code=400)

        # Handle direct Gemini or DeepSeek model selection
        # When user selects these models, bypass the GPT agent and call APIs directly
        if selected_model == "gemini-2.5-flash":
            logger.info(f"🔮 [DIRECT-GEMINI] Using Gemini 2.5 Flash directly")

            async def generate_gemini():
                try:
                    # Using stable version for production reliability
                    model = genai.GenerativeModel('gemini-2.5-flash')
                    response = await model.generate_content_async(
                        user_message,
                        generation_config=genai.types.GenerationConfig(
                            max_output_tokens=4096,
                            temperature=0.7,
                        ),
                        stream=True
                    )

                    accumulated_text = ""
                    async for chunk in response:
                        if chunk.text:
                            accumulated_text += chunk.text
                            event_dict = {
                                "type": "text_delta",
                                "delta": {"text": chunk.text}
                            }
                            yield f"data: {json.dumps(event_dict, ensure_ascii=False)}\n\n"

                    # Save to Firebase in background
                    if user_id and accumulated_text:
                        async def save_gemini():
                            try:
                                firebase = get_firebase_service()
                                chat_id = conversation_id or f"gemini_{user_id}_{int(datetime.now().timestamp())}"
                                await firebase.save_chat_messages_batch(
                                    user_id=user_id,
                                    chat_id=chat_id,
                                    user_message=user_message,
                                    assistant_message=accumulated_text,
                                    model="gemini-2.5-flash"
                                )
                                await firebase.update_user_limits(user_id, "premium")
                            except Exception as e:
                                logger.error(f"Firebase error: {e}")
                        asyncio.create_task(save_gemini())

                    yield "data: [DONE]\n\n"

                except Exception as e:
                    logger.error(f"❌ [DIRECT-GEMINI] Error: {e}", exc_info=True)
                    error_event = {
                        "type": "text_delta",
                        "delta": {"text": f"Ошибка Gemini: {str(e)}"}
                    }
                    yield f"data: {json.dumps(error_event, ensure_ascii=False)}\n\n"
                    yield "data: [DONE]\n\n"

            return StreamingResponse(generate_gemini(), media_type="text/event-stream")

        elif selected_model == "deepseek-chat":
            logger.info(f"🔮 [DIRECT-DEEPSEEK] Using DeepSeek V3 directly")

            async def generate_deepseek():
                try:
                    async with httpx.AsyncClient(timeout=120.0) as client:
                        async with client.stream(
                            "POST",
                            "https://api.deepseek.com/v1/chat/completions",
                            headers={
                                "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
                                "Content-Type": "application/json"
                            },
                            json={
                                "model": "deepseek-chat",
                                "messages": [{"role": "user", "content": user_message}],
                                "max_tokens": 4096,
                                "temperature": 0.7,
                                "stream": True
                            }
                        ) as response:
                            if response.status_code != 200:
                                error_text = await response.aread()
                                error_event = {
                                    "type": "text_delta",
                                    "delta": {"text": f"Ошибка DeepSeek API: {response.status_code}"}
                                }
                                yield f"data: {json.dumps(error_event, ensure_ascii=False)}\n\n"
                                yield "data: [DONE]\n\n"
                                return

                            accumulated_text = ""
                            async for line in response.aiter_lines():
                                if line.startswith("data: "):
                                    data_str = line[6:]
                                    if data_str == "[DONE]":
                                        break
                                    try:
                                        data = json.loads(data_str)
                                        if "choices" in data and len(data["choices"]) > 0:
                                            delta = data["choices"][0].get("delta", {})
                                            content = delta.get("content", "")
                                            if content:
                                                accumulated_text += content
                                                event_dict = {
                                                    "type": "text_delta",
                                                    "delta": {"text": content}
                                                }
                                                yield f"data: {json.dumps(event_dict, ensure_ascii=False)}\n\n"
                                    except json.JSONDecodeError:
                                        pass

                            # Save to Firebase in background
                            if user_id and accumulated_text:
                                async def save_deepseek():
                                    try:
                                        firebase = get_firebase_service()
                                        chat_id = conversation_id or f"deepseek_{user_id}_{int(datetime.now().timestamp())}"
                                        await firebase.save_chat_messages_batch(
                                            user_id=user_id,
                                            chat_id=chat_id,
                                            user_message=user_message,
                                            assistant_message=accumulated_text,
                                            model="deepseek-v3"
                                        )
                                        await firebase.update_user_limits(user_id, "premium")
                                    except Exception as e:
                                        logger.error(f"Firebase error: {e}")
                                asyncio.create_task(save_deepseek())

                            yield "data: [DONE]\n\n"

                except Exception as e:
                    logger.error(f"❌ [DIRECT-DEEPSEEK] Error: {e}", exc_info=True)
                    error_event = {
                        "type": "text_delta",
                        "delta": {"text": f"Ошибка DeepSeek: {str(e)}"}
                    }
                    yield f"data: {json.dumps(error_event, ensure_ascii=False)}\n\n"
                    yield "data: [DONE]\n\n"

            return StreamingResponse(generate_deepseek(), media_type="text/event-stream")

        # Default: use GPT-5 agent
        # Create simple context object for widget support (no thread/store needed)
        class SimpleContext:
            def __init__(self):
                self._custom_widgets = []

        agent_context = SimpleContext()

        # Create OpenAI Conversations Session for persistent chat history
        # If conversation_id is provided, resume that conversation
        # Otherwise, create a new conversation
        if conversation_id:
            logger.info(f"📝 Resuming conversation: {conversation_id}")
            session = OpenAIConversationsSession(conversation_id=conversation_id)
        else:
            logger.info(f"📝 Creating new conversation session")
            session = OpenAIConversationsSession()

        # Determine if this request has files - if so, don't provide weather/AQI tools
        has_files = len(uploaded_files) > 0

        # Create agent with tools - use gpt-5-mini
        # IMPORTANT: Remove weather/AQI tools when processing files to prevent false triggers
        if has_files:
            agent_tools = [web_search_tool, create_document, send_reaction_gif]  # Web search, document creation, and GIFs for file requests
            tools_note = "You have web_search_tool, create_document, and send_reaction_gif available. Weather and AQI tools are disabled for file processing."
        else:
            agent_tools = [web_search_tool, get_weather_simple, get_aqi_simple, ask_gemini, ask_deepseek, create_document, send_reaction_gif]
            tools_note = """Tools available:
- web_search_tool: For current information, news, and facts
- get_weather_simple: ONLY for explicit weather requests (user must say: погода, weather, температура, forecast)
- get_aqi_simple: ONLY for explicit air quality requests (user must say: качество воздуха, AQI, загрязнение)
- ask_gemini: Handoff to Gemini 2.5 Flash when user asks for "gemini" or needs advanced reasoning
- ask_deepseek: Handoff to DeepSeek V3 when user asks for "deepseek" or needs complex coding help
- create_document: Create presentations (PDF), Excel tables, or Word documents
- send_reaction_gif: Send a reaction GIF to make conversation engaging (use sparingly!)"""

        agent = Agent(
            name="Custom ChatKit Assistant",
            model="gpt-5-mini",
            instructions=f"""You are a helpful AI assistant.
You have persistent memory - you remember all previous messages in this conversation.

Always respond in the same language the user asks in.
Be polite, informative, and concise.

EMOJI USAGE (IMPORTANT):
- Use relevant emojis in your responses to make them more engaging and friendly
- Add emojis at the start of sections or important points: 📌 🎯 💡 ✨ 🔥 ⭐ 📝 🚀 ✅ ❌ ⚠️ 💪 👍
- Use emojis for lists: 1️⃣ 2️⃣ 3️⃣ or • with emoji prefix
- Weather: 🌤️ ☀️ 🌧️ ❄️ 🌡️ | Food: 🍕 🍔 🍜 | Time: ⏰ 📅 | Money: 💰 💵
- Emotions: 😊 😄 🤔 😮 | Work: 💼 📊 📈 | Tech: 💻 📱 🔧
- Example: "📌 **Главные пункты:**\n\n1️⃣ Первый пункт\n2️⃣ Второй пункт"
- Don't overuse - 2-4 emojis per message is optimal

MARKDOWN FORMATTING (CRITICAL):
- ALWAYS use proper markdown formatting with line breaks
- For numbered lists: put EACH item on a NEW LINE with empty line before the list
- For bullet lists: put EACH item on a NEW LINE with empty line before the list
- Example of CORRECT numbered list format:

📋 Here is the plan:

1️⃣ First item - description
2️⃣ Second item - description
3️⃣ Third item - description

- NEVER write lists inline like "text1.item text2.item" - ALWAYS use line breaks
- Use **bold** for emphasis, headers with ## when appropriate

CRITICAL: After using ANY tool (including web_search), you MUST ALWAYS provide a final text response to the user summarizing what you found. NEVER end your turn without a text response.

DOCUMENT/IMAGE PROCESSING:
- When messages contain "📄 **filename**:" sections, these are FILE ANALYSIS RESULTS
- The text after the filename is the analysis of that file's contents
- Just discuss and answer questions about the file based on the analysis provided
- DO NOT call any tools when processing files

{tools_note}

AI MODEL HANDOFF:
- Use ask_gemini when user explicitly asks for "Gemini", "гемини", or needs advanced multi-step reasoning
- Use ask_deepseek when user explicitly asks for "DeepSeek", "дипсик", or needs complex coding/technical help
- CRITICAL: When you call ask_gemini or ask_deepseek, these tools will STREAM their response directly to the user
- After calling ask_gemini or ask_deepseek, DO NOT add any additional text - the tool output IS the final response
- The streaming output from these tools already includes proper attribution (e.g., "Ответ от Gemini 2.5 Flash:")
- DO NOT handoff unless user explicitly requests it or the task clearly requires specialized model capabilities

DOCUMENT CREATION (IMPORTANT):
- Use create_document tool when user asks to create presentations, Excel tables, or Word documents
- Keywords: создай презентацию, сделай таблицу, напиши документ, договор, отчёт, excel, word, ppt
- When user confirms with "можно", "да", "начинай", "делай" - call create_document immediately
- Pass the FULL task description from conversation context, not just user's last message
- After calling create_document, provide a brief confirmation message

REACTION GIFS (IMPORTANT - Makes conversation feel human-like):
- Use send_reaction_gif to send GIFs that make the chat feel like talking to a friend
- WHEN TO USE (sparingly, max 1 per 3-4 messages):
  • Greetings: "привет", "здравствуй", "hi" → mood: "greeting"
  • Celebrations: good news, achievements, success → mood: "celebration" or "applause"
  • Agreement/confirmation: "хорошо", "отлично" → mood: "thumbs_up"
  • Jokes/humor: when something funny → mood: "funny"
  • Thinking: complex questions → mood: "thinking"
  • Exciting news: → mood: "excited"
- WHEN NOT TO USE:
  • Technical/serious questions
  • Every single message
  • When user seems busy or wants quick answer
- Call send_reaction_gif BEFORE your text response to show GIF first

STRICT RULES:
- NEVER use file names, company names, or document titles as location parameters
- Weather/AQI tools need EXPLICIT user request with keywords like "погода", "weather", "AQI", "качество воздуха"
- If user asks about a document/file, just respond based on the analysis - NO tool calls
- EXCEPTION: When you call ask_gemini or ask_deepseek, DO NOT add any text after - their output is already complete
""",
            tools=agent_tools,
        )

        # Process all files with Claude Haiku (supports both images and documents)
        # OPTIMIZED: Process files in parallel for faster response
        file_analyses = []
        if uploaded_files:
            claude_agent = get_claude_agent()
            task = user_message if user_message else "Проанализируй этот файл и предоставь подробное описание содержимого."

            async def process_single_file(file_info):
                """Process a single file and return result."""
                try:
                    logger.info(f"Processing file with Claude Haiku: {file_info['filename']} ({file_info['content_type']})")
                    analysis = await claude_agent.process_document(
                        file_bytes=file_info['bytes'],
                        file_name=file_info['filename'],
                        mime_type=file_info['content_type'],
                        task=task
                    )
                    logger.info(f"✅ File processed successfully: {file_info['filename']}")
                    return {"filename": file_info['filename'], "analysis": analysis}
                except Exception as e:
                    logger.error(f"Error processing file {file_info['filename']}: {e}")
                    return {"filename": file_info['filename'], "analysis": f"Ошибка при обработке: {str(e)}"}

            # Process all files in parallel
            if len(uploaded_files) == 1:
                # Single file - process directly
                file_analyses = [await process_single_file(uploaded_files[0])]
            else:
                # Multiple files - process in parallel
                logger.info(f"🚀 Processing {len(uploaded_files)} files in parallel...")
                tasks = [process_single_file(f) for f in uploaded_files]
                file_analyses = await asyncio.gather(*tasks)
                logger.info(f"✅ All {len(uploaded_files)} files processed in parallel")

        # Build input message with file analyses
        input_parts = []

        # Add file analyses
        if file_analyses:
            for fa in file_analyses:
                input_parts.append(f"📄 **{fa['filename']}**:\n{fa['analysis']}")

        # Add user message
        if user_message:
            input_parts.append(user_message)
        elif not input_parts:
            input_parts.append("Проанализируй загруженные файлы.")

        input_message = "\n\n---\n\n".join(input_parts)

        logger.info(f"🚀 [CUSTOM-CHAT] Agent run starting with message length: {len(input_message)}")

        # Stream response as SSE
        async def generate():
            # Helper function to extract sources
            def extract_sources_from_obj(obj, sources_list):
                if obj is None: return
                if isinstance(obj, dict):
                    if obj.get('type') == 'url_citation' and obj.get('url'):
                        source = {"url": obj.get('url'), "title": obj.get('title', '') or obj.get('text', '')}
                        if source not in sources_list: sources_list.append(source)
                    for value in obj.values(): extract_sources_from_obj(value, sources_list)
                elif isinstance(obj, list):
                    for item in obj: extract_sources_from_obj(item, sources_list)
                elif hasattr(obj, '__dict__'):
                    obj_type = getattr(obj, 'type', None)
                    if str(obj_type) == 'url_citation':
                        url = getattr(obj, 'url', None)
                        if url:
                            source = {"url": url, "title": getattr(obj, 'title', '') or getattr(obj, 'text', '')}
                            if source not in sources_list: sources_list.append(source)
                    for attr in ['annotations', 'content', 'output']:
                        val = getattr(obj, attr, None)
                        if val: extract_sources_from_obj(val, sources_list)

            max_retries = 3
            retry_count = 0

            while retry_count < max_retries:
                try:
                    logger.info(f"Starting streaming response... (attempt {retry_count + 1}/{max_retries})")
                    collected_sources = []
                    accumulated_text = ""

                    # Run agent with streaming
                    result = Runner.run_streamed(
                        agent,
                        input_message,
                        context=agent_context,
                        session=session,
                        run_config=RunConfig(
                            model_settings=ModelSettings(
                                max_tokens=2048,
                            ),
                        ),
                        max_turns=30,
                    )

                    event_counter = 0
                    sent_widget_count = 0


                    async for event in result.stream_events():
                        event_counter += 1
                        event_type = getattr(event, 'type', None)
                        logger.info(f"📦 [EVENT {event_counter}] Type: {event_type}")

                        if event_type and 'reasoning' in str(event_type).lower():
                            continue

                        # Handle tool output streaming
                        if event_type == "run_item_stream_event":
                            if hasattr(event, 'item'):
                                item = event.item
                                item_type = getattr(item, 'type', None)
                                if item_type == "tool_call_output_item":
                                    output = getattr(item, 'output', None)
                                    if output and isinstance(output, str):
                                        logger.info(f"🔧 [TOOL OUTPUT] {output[:50]}...")
                                        # Skip GIF tool output - it's handled separately via reaction_gif event
                                        if "[GIF отправлен:" in output:
                                            logger.info(f"🎬 [SKIP] Skipping GIF tool output from text stream")
                                            continue
                                        # Skip document creation tool output
                                        if "Начинаю создание:" in output:
                                            logger.info(f"📄 [SKIP] Skipping document tool output from text stream")
                                            continue
                                        accumulated_text += output
                                        event_dict = {
                                            "type": "text_delta",
                                            "delta": {"text": output}
                                        }
                                        yield f"data: {json.dumps(event_dict, ensure_ascii=False)}\n\n"
                                        # yield control to event loop for smooth streaming
                                        await asyncio.sleep(0)
                                        continue

                        # Main text content streaming
                        if hasattr(event, 'data'):
                            response_data = event.data
                            
                            # Skip function calls info
                            if hasattr(response_data, 'type'):
                                resp_type = str(response_data.type)
                                if 'function_call' in resp_type.lower() or 'tool_call' in resp_type.lower():
                                    continue

                            if hasattr(response_data, 'delta'):
                                delta = response_data.delta
                                text_content = None

                                if isinstance(delta, str) and delta:
                                    text_content = delta
                                elif hasattr(delta, 'text') and delta.text:
                                    text_content = delta.text

                                if text_content:
                                    # Очистка от JSON location (оставляем, если это специфика ваших тулов)
                                    text_content = re.sub(r'\{"location"\s*:\s*"[^"]*"\}', '', text_content)
                                    text_content = re.sub(r"\{'location'\s*:\s*'[^']*'\}", '', text_content)

                                    accumulated_text += text_content

                                    # Отправляем СРАЗУ, без буфера
                                    event_dict = {
                                        "type": "text_delta",
                                        "delta": {"text": text_content}
                                    }
                                    logger.info(f"📤 [YIELD] Sending text chunk: {text_content[:50]}...")
                                    yield f"data: {json.dumps(event_dict, ensure_ascii=False)}\n\n"
                                    await asyncio.sleep(0) # Даем шанс Event Loop отправить пакет

                        # Check for new widgets
                        if hasattr(agent_context, '_custom_widgets') and agent_context._custom_widgets:
                            new_widgets = agent_context._custom_widgets[sent_widget_count:]
                            for widget_data in new_widgets:
                                widget_event = {
                                    "type": "widget",
                                    "widget": widget_data
                                }
                                yield f"data: {json.dumps(widget_event, ensure_ascii=False)}\n\n"
                                await asyncio.sleep(0)
                                sent_widget_count += 1

                        # Check for document creation request
                        if hasattr(agent_context, '_document_request') and agent_context._document_request:
                            doc_request = agent_context._document_request
                            agent_context._document_request = None  # Reset after sending
                            logger.info(f"📄 [DOCUMENT] Sending document request: {doc_request}")
                            doc_event = {
                                "type": "create_document",
                                "document_type": doc_request.get("_document_type"),
                                "task": doc_request.get("task"),
                                "title": doc_request.get("title")
                            }
                            yield f"data: {json.dumps(doc_event, ensure_ascii=False)}\n\n"
                            await asyncio.sleep(0)

                        # Check for reaction GIF
                        if hasattr(agent_context, '_reaction_gif') and agent_context._reaction_gif:
                            gif_data = agent_context._reaction_gif
                            agent_context._reaction_gif = None  # Reset after sending
                            logger.info(f"🎬 [GIF] Sending reaction GIF: {gif_data}")
                            gif_event = {
                                "type": "reaction_gif",
                                "url": gif_data.get("url"),
                                "mood": gif_data.get("mood")
                            }
                            yield f"data: {json.dumps(gif_event, ensure_ascii=False)}\n\n"
                            await asyncio.sleep(0)

                        # Extract sources
                        extract_sources_from_obj(event, collected_sources)

                    # ⬇️ Код ниже ПОСЛЕ цикла async for ⬇️
                    logger.info(f"✅ [STREAM END] Processed {event_counter} events, accumulated {len(accumulated_text)} chars")

                    # Send sources if collected
                    if collected_sources:
                        sources_event = {
                            "type": "sources",
                            "sources": collected_sources
                        }
                        yield f"data: {json.dumps(sources_event, ensure_ascii=False)}\n\n"

                    # Send conversation_id
                    new_conversation_id = getattr(session, '_session_id', None)
                    if new_conversation_id:
                        conv_event = {
                            "type": "conversation_id",
                            "conversation_id": new_conversation_id
                        }
                        yield f"data: {json.dumps(conv_event, ensure_ascii=False)}\n\n"

                    # Save messages to Firestore and update limits (in background)
                    if user_id and accumulated_text:
                        async def save_to_firebase():
                            try:
                                firebase = get_firebase_service()
                                # Use conversation_id as chat_id, or generate one
                                chat_id = new_conversation_id or conversation_id or f"chat_{user_id}_{int(datetime.now().timestamp())}"

                                # Save both messages in batch
                                await firebase.save_chat_messages_batch(
                                    user_id=user_id,
                                    chat_id=chat_id,
                                    user_message=user_message,
                                    assistant_message=accumulated_text,
                                    model=selected_model,
                                    sources=collected_sources if collected_sources else None
                                )

                                # Update limits
                                await firebase.update_user_limits(user_id, "free")
                            except Exception as save_error:
                                logger.error(f"Failed to save to Firebase: {save_error}")

                        # Run in background to not block [DONE]
                        asyncio.create_task(save_to_firebase())

                    yield "data: [DONE]\n\n"
                    return  # Успешное завершение стрима

                except ModelBehaviorError as e:
                    # Retry logic
                    retry_count += 1
                    logger.warning(f"ModelBehaviorError: {e}")
                    if retry_count < max_retries:
                        await asyncio.sleep(0.5)
                        if hasattr(agent_context, '_custom_widgets'):
                            agent_context._custom_widgets = []
                        continue
                    yield "data: [DONE]\n\n"
                    return

                except Exception as e:
                    # Handle other errors
                    logger.error(f"Stream error: {e}", exc_info=True)
                    yield "data: [DONE]\n\n"
                    return

        return StreamingResponse(generate(), media_type="text/event-stream")

    except Exception as e:
        logger.error(f"Error in custom_chat_endpoint: {e}", exc_info=True)
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/api/aqi-widget")
async def aqi_widget_endpoint():
    """Get AQI widget data for Tashkent using IQAir API."""
    try:
        client = get_http_client()
        # Get coordinates for Tashkent
        geo_response = await client.get(
            "https://geocoding-api.open-meteo.com/v1/search",
            params={"name": "Ташкент", "count": 1, "language": "ru", "format": "json"},
            timeout=10.0
        )

        if geo_response.status_code != 200:
            return JSONResponse({"error": "Could not find Tashkent"}, status_code=500)

        geo_data = geo_response.json()
        if not geo_data.get("results"):
            return JSONResponse({"error": "Tashkent not found"}, status_code=500)

        result = geo_data["results"][0]
        lat = result["latitude"]
        lon = result["longitude"]

        # Get air quality data from IQAir API
        iqair_response = await client.get(
            "https://api.airvisual.com/v2/nearest_city",
            params={
                "lat": lat,
                "lon": lon,
                "key": IQAIR_API_KEY
            },
            timeout=15.0
        )

        if iqair_response.status_code != 200:
            logger.error(f"IQAir API error: {iqair_response.status_code} - {iqair_response.text}")
            return JSONResponse({"error": "Could not fetch air quality data from IQAir"}, status_code=500)

        iqair_data = iqair_response.json()

        if iqair_data.get("status") != "success":
            logger.error(f"IQAir API returned error: {iqair_data}")
            return JSONResponse({"error": "IQAir API error"}, status_code=500)

        data = iqair_data.get("data", {})
        current = data.get("current", {})
        pollution = current.get("pollution", {})
        weather = current.get("weather", {})

        # Get AQI value (US AQI standard)
        aqi_value = int(pollution.get("aqius", 0))

        # Determine category based on US AQI
        if aqi_value <= 50:
            category = "Хорошо"
        elif aqi_value <= 100:
            category = "Умеренно"
        elif aqi_value <= 150:
            category = "Вредно для чувствительных"
        elif aqi_value <= 200:
            category = "Вредно"
        elif aqi_value <= 300:
            category = "Очень вредно"
        else:
            category = "Опасно"

        # Map pollutant code to name
        pollutant_map = {
            "p2": "PM2.5",
            "p1": "PM10",
            "o3": "O3",
            "n2": "NO2",
            "s2": "SO2",
            "co": "CO"
        }
        main_pollutant = pollution.get("mainus", "p2")
        pollutant = pollutant_map.get(main_pollutant, "PM2.5")

        # Get current time
        now = datetime.now()
        updated = now.strftime("%H:%M, %d %B")

        # Get real weather data from IQAir
        temp = f"{round(weather.get('tp', 0))}°C"
        humidity = f"{round(weather.get('hu', 0))}%"
        wind_speed_ms = weather.get('ws', 0)
        wind = f"{round(wind_speed_ms * 3.6)} км/ч"

        # Return widget data as JSON
        return JSONResponse({
            "city": data.get("city", "Ташкент"),
            "aqi": aqi_value,
            "category": category,
            "scale": "US AQI",
            "pollutant": pollutant,
            "temp": temp,
            "humidity": humidity,
            "wind": wind,
            "updated": updated
        })

    except Exception as e:
        logger.error(f"AQI widget endpoint error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
# Document Generation Endpoints
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/create-document")
async def create_document_endpoint(request: Request):
    """
    Create documents (PDF, Excel, Word, Presentations) with streaming.
    Supports:
    - presentation: HTML -> PDF presentation
    - excel: Excel spreadsheet
    - word: Word document

    Returns SSE stream with progress updates and final file download link.
    """
    logger.info("📄 [DOCUMENT-CREATE] New document creation request")

    try:
        content_type = request.headers.get("content-type", "")

        if "multipart/form-data" in content_type:
            form = await request.form()
            doc_type = form.get("type", "word")  # presentation, excel, word
            task = form.get("task", "")
            user_id = form.get("user_id")
            language = form.get("language", "ru")

            # Check for file to modify
            uploaded_file = form.get("file")
            file_bytes = None
            file_name = None
            if uploaded_file and hasattr(uploaded_file, 'file'):
                file_bytes = await uploaded_file.read()
                file_name = uploaded_file.filename
        else:
            data = await request.json()
            doc_type = data.get("type", "word")
            task = data.get("task", "")
            user_id = data.get("user_id")
            language = data.get("language", "ru")
            file_bytes = None
            file_name = None

            # Check for base64 encoded file
            if data.get("file_base64"):
                file_bytes = base64.b64decode(data["file_base64"])
                file_name = data.get("file_name", "document")

        logger.info(f"📄 [DOCUMENT-CREATE] Type: {doc_type}, Task: {task[:100]}...")

        if not task:
            return JSONResponse({"error": "Task is required"}, status_code=400)

        claude_agent = get_claude_agent()

        async def generate_document():
            """Stream document creation progress."""
            try:
                document_id = f"doc_{int(asyncio.get_event_loop().time() * 1000)}"
                final_bytes = None
                final_filename = None

                if doc_type == "presentation":
                    # Create presentation
                    final_filename = "presentation.pdf"
                    html_content = ""

                    async for event_type, content in claude_agent.create_presentation_html(
                        topic=task,
                        requirements="",
                        language=language
                    ):
                        if event_type == "html_chunk":
                            # Stream HTML for preview
                            event = {
                                "type": "document_chunk",
                                "chunk_type": "html",
                                "content": content
                            }
                            yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"
                            html_content = content if not html_content else html_content
                        elif event_type == "complete":
                            html_content = content
                            # Convert to PDF
                            status_event = {"type": "status", "message": "Конвертирую в PDF..."}
                            yield f"data: {json.dumps(status_event, ensure_ascii=False)}\n\n"
                            try:
                                final_bytes = await claude_agent.html_to_pdf(html_content)
                            except ImportError:
                                # If WeasyPrint not available, return HTML
                                final_bytes = html_content.encode('utf-8')
                                final_filename = "presentation.html"
                        elif event_type == "error":
                            error_event = {"type": "error", "message": content}
                            yield f"data: {json.dumps(error_event, ensure_ascii=False)}\n\n"
                            yield "data: [DONE]\n\n"
                            return

                elif doc_type == "excel":
                    final_filename = "document.xlsx"

                    if file_bytes and file_name:
                        # Modify existing Excel
                        async for event_type, content in claude_agent.modify_excel_document(
                            file_bytes=file_bytes,
                            file_name=file_name,
                            instructions=task,
                            language=language
                        ):
                            if event_type == "status":
                                status_event = {"type": "status", "message": content}
                                yield f"data: {json.dumps(status_event, ensure_ascii=False)}\n\n"
                            elif event_type == "data":
                                data_event = {"type": "data", "content": content}
                                yield f"data: {json.dumps(data_event, ensure_ascii=False)}\n\n"
                            elif event_type == "complete":
                                final_bytes = content
                                final_filename = file_name
                            elif event_type == "error":
                                error_event = {"type": "error", "message": content}
                                yield f"data: {json.dumps(error_event, ensure_ascii=False)}\n\n"
                                yield "data: [DONE]\n\n"
                                return
                    else:
                        # Create new Excel
                        async for event_type, content in claude_agent.create_excel_document(
                            task=task,
                            language=language
                        ):
                            if event_type == "status":
                                status_event = {"type": "status", "message": content}
                                yield f"data: {json.dumps(status_event, ensure_ascii=False)}\n\n"
                            elif event_type == "data":
                                data_event = {"type": "data", "content": content}
                                yield f"data: {json.dumps(data_event, ensure_ascii=False)}\n\n"
                            elif event_type == "complete":
                                final_bytes = content
                            elif event_type == "error":
                                error_event = {"type": "error", "message": content}
                                yield f"data: {json.dumps(error_event, ensure_ascii=False)}\n\n"
                                yield "data: [DONE]\n\n"
                                return

                elif doc_type == "word":
                    final_filename = "document.docx"

                    if file_bytes and file_name:
                        # Modify existing Word
                        async for event_type, content in claude_agent.modify_word_document(
                            file_bytes=file_bytes,
                            file_name=file_name,
                            instructions=task,
                            language=language
                        ):
                            if event_type == "text_chunk":
                                # Stream text for preview
                                event = {
                                    "type": "document_chunk",
                                    "chunk_type": "text",
                                    "content": content
                                }
                                yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"
                            elif event_type == "status":
                                status_event = {"type": "status", "message": content}
                                yield f"data: {json.dumps(status_event, ensure_ascii=False)}\n\n"
                            elif event_type == "complete":
                                final_bytes = content
                                final_filename = file_name
                            elif event_type == "error":
                                error_event = {"type": "error", "message": content}
                                yield f"data: {json.dumps(error_event, ensure_ascii=False)}\n\n"
                                yield "data: [DONE]\n\n"
                                return
                    else:
                        # Create new Word
                        async for event_type, content in claude_agent.create_word_document(
                            task=task,
                            language=language
                        ):
                            if event_type == "text_chunk":
                                event = {
                                    "type": "document_chunk",
                                    "chunk_type": "text",
                                    "content": content
                                }
                                yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"
                            elif event_type == "status":
                                status_event = {"type": "status", "message": content}
                                yield f"data: {json.dumps(status_event, ensure_ascii=False)}\n\n"
                            elif event_type == "complete":
                                final_bytes = content
                            elif event_type == "error":
                                error_event = {"type": "error", "message": content}
                                yield f"data: {json.dumps(error_event, ensure_ascii=False)}\n\n"
                                yield "data: [DONE]\n\n"
                                return

                else:
                    error_event = {"type": "error", "message": f"Unknown document type: {doc_type}"}
                    yield f"data: {json.dumps(error_event, ensure_ascii=False)}\n\n"
                    yield "data: [DONE]\n\n"
                    return

                # Save file and return download link
                if final_bytes:
                    # Save to workspace
                    import uuid
                    file_id = str(uuid.uuid4())[:8]
                    save_filename = f"{file_id}_{final_filename}"
                    save_path = os.path.join(CLAUDE_WORKSPACE_DIR, save_filename)

                    with open(save_path, "wb") as f:
                        f.write(final_bytes)

                    logger.info(f"✅ [DOCUMENT-CREATE] Saved: {save_path}")

                    # Return completion with download info
                    complete_event = {
                        "type": "complete",
                        "document_id": document_id,
                        "filename": final_filename,
                        "file_id": file_id,
                        "size": len(final_bytes),
                        "download_url": f"/api/download-document/{file_id}/{final_filename}",
                        "file_base64": base64.b64encode(final_bytes).decode('utf-8')
                    }
                    yield f"data: {json.dumps(complete_event, ensure_ascii=False)}\n\n"

                yield "data: [DONE]\n\n"

            except Exception as e:
                logger.error(f"❌ [DOCUMENT-CREATE] Error: {e}", exc_info=True)
                error_event = {"type": "error", "message": str(e)}
                yield f"data: {json.dumps(error_event, ensure_ascii=False)}\n\n"
                yield "data: [DONE]\n\n"

        return StreamingResponse(generate_document(), media_type="text/event-stream")

    except Exception as e:
        logger.error(f"Error in create_document_endpoint: {e}", exc_info=True)
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/api/download-document/{file_id}/{filename}")
async def download_document_endpoint(file_id: str, filename: str):
    """Download a generated document."""
    try:
        save_filename = f"{file_id}_{filename}"
        file_path = os.path.join(CLAUDE_WORKSPACE_DIR, save_filename)

        if not os.path.exists(file_path):
            return JSONResponse({"error": "File not found"}, status_code=404)

        from fastapi.responses import FileResponse

        # Determine media type
        if filename.endswith(".pdf"):
            media_type = "application/pdf"
        elif filename.endswith(".xlsx"):
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif filename.endswith(".docx"):
            media_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        elif filename.endswith(".html"):
            media_type = "text/html"
        else:
            media_type = "application/octet-stream"

        return FileResponse(
            path=file_path,
            filename=filename,
            media_type=media_type
        )

    except Exception as e:
        logger.error(f"Download error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/api/convert-to-html")
async def convert_to_html_endpoint(request: Request):
    """
    Convert Word (.docx) or Excel (.xlsx) file to HTML for preview.
    Accepts base64 encoded file content.
    """
    try:
        data = await request.json()
        file_base64 = data.get("file_base64")
        filename = data.get("filename", "document")

        if not file_base64:
            return JSONResponse({"error": "file_base64 is required"}, status_code=400)

        # Decode base64
        file_bytes = base64.b64decode(file_base64)

        html_content = ""

        if filename.endswith(".docx"):
            # Convert Word to HTML using python-docx and mammoth
            try:
                import mammoth
                import io

                result = mammoth.convert_to_html(io.BytesIO(file_bytes))
                html_content = result.value

                # Wrap in styled HTML
                html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <style>
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif;
            max-width: 800px;
            margin: 0 auto;
            padding: 40px 20px;
            line-height: 1.6;
            color: #1a1a1a;
            background: #fff;
        }}
        h1 {{ font-size: 2em; margin: 0.5em 0; border-bottom: 1px solid #eee; padding-bottom: 0.3em; }}
        h2 {{ font-size: 1.5em; margin: 0.5em 0; }}
        h3 {{ font-size: 1.25em; margin: 0.5em 0; }}
        p {{ margin: 0.5em 0; }}
        ul, ol {{ margin: 0.5em 0; padding-left: 2em; }}
        li {{ margin: 0.25em 0; }}
        table {{ border-collapse: collapse; width: 100%; margin: 1em 0; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #f5f5f5; font-weight: 600; }}
        img {{ max-width: 100%; height: auto; }}
        strong {{ font-weight: 600; }}
        em {{ font-style: italic; }}
    </style>
</head>
<body>
{html_content}
</body>
</html>"""

            except ImportError:
                # Fallback to basic python-docx if mammoth not available
                from docx import Document
                import io

                doc = Document(io.BytesIO(file_bytes))
                paragraphs_html = []

                for para in doc.paragraphs:
                    if para.style.name.startswith('Heading'):
                        level = para.style.name[-1] if para.style.name[-1].isdigit() else '1'
                        paragraphs_html.append(f"<h{level}>{para.text}</h{level}>")
                    elif para.text.strip():
                        paragraphs_html.append(f"<p>{para.text}</p>")

                html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        body {{ font-family: -apple-system, BlinkMacSystemFont, sans-serif; max-width: 800px; margin: 0 auto; padding: 40px 20px; line-height: 1.6; }}
        h1, h2, h3 {{ margin: 0.5em 0; }}
        p {{ margin: 0.5em 0; }}
    </style>
</head>
<body>
{''.join(paragraphs_html)}
</body>
</html>"""

        elif filename.endswith(".xlsx"):
            # Convert Excel to HTML using openpyxl
            from openpyxl import load_workbook
            import io

            wb = load_workbook(io.BytesIO(file_bytes))
            sheets_html = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Build table HTML
                table_rows = []
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    cells = []
                    for cell_value in row:
                        value = str(cell_value) if cell_value is not None else ""
                        if row_idx == 1:
                            cells.append(f"<th>{value}</th>")
                        else:
                            cells.append(f"<td>{value}</td>")
                    if any(cells):
                        tag = "thead" if row_idx == 1 else "tbody"
                        if row_idx == 1:
                            table_rows.append(f"<tr>{''.join(cells)}</tr>")
                        else:
                            table_rows.append(f"<tr>{''.join(cells)}</tr>")

                if table_rows:
                    # Separate header from body
                    header = f"<thead>{table_rows[0]}</thead>" if table_rows else ""
                    body = f"<tbody>{''.join(table_rows[1:])}</tbody>" if len(table_rows) > 1 else ""

                    sheets_html.append(f"""
                    <div class="sheet">
                        <h2>{sheet_name}</h2>
                        <div class="table-container">
                            <table>
                                {header}
                                {body}
                            </table>
                        </div>
                    </div>
                    """)

            html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <style>
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            margin: 0;
            padding: 20px;
            background: #f5f5f5;
            color: #1a1a1a;
        }}
        .sheet {{
            background: #fff;
            border-radius: 8px;
            padding: 20px;
            margin-bottom: 20px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        }}
        .sheet h2 {{
            margin: 0 0 16px 0;
            font-size: 18px;
            color: #22c55e;
        }}
        .table-container {{
            overflow-x: auto;
        }}
        table {{
            border-collapse: collapse;
            width: 100%;
            font-size: 14px;
        }}
        th, td {{
            border: 1px solid #e5e7eb;
            padding: 10px 12px;
            text-align: left;
            white-space: nowrap;
        }}
        th {{
            background: #22c55e;
            color: white;
            font-weight: 600;
            position: sticky;
            top: 0;
        }}
        tr:nth-child(even) {{
            background: #f9fafb;
        }}
        tr:hover {{
            background: #f0fdf4;
        }}
    </style>
</head>
<body>
{''.join(sheets_html)}
</body>
</html>"""

        elif filename.endswith(".pdf"):
            # For PDF, we can't easily convert to HTML, return error
            return JSONResponse({
                "error": "PDF files are displayed natively in the browser",
                "use_native": True
            }, status_code=200)

        else:
            return JSONResponse({"error": f"Unsupported file type: {filename}"}, status_code=400)

        # Return HTML content as base64
        html_base64 = base64.b64encode(html_content.encode('utf-8')).decode('utf-8')

        return JSONResponse({
            "html_base64": html_base64,
            "html_length": len(html_content)
        })

    except Exception as e:
        logger.error(f"Convert to HTML error: {e}", exc_info=True)
        return JSONResponse({"error": str(e)}, status_code=500)


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    print(f"Starting ChatKit backend on port {port}")
    uvicorn.run(app, host="0.0.0.0", port=port)
